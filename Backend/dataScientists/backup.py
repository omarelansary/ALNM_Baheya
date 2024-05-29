from django.shortcuts import render
from django.http import HttpResponse
#edited
from rest_framework.decorators import api_view 
from rest_framework.response import Response
from rest_framework import status 
#Allows us to fetch and create users and tokens from our database
from rest_framework.authtoken.models import Token 
from django.contrib.auth.models import User
from django. shortcuts import get_object_or_404
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from .models import DataScientist
from django.contrib.auth.hashers import make_password
from django.core.serializers import serialize
import bcrypt
import jwt
from datetime import datetime, timedelta
from django.conf import settings
from django.contrib.auth.hashers import check_password
from django.core.mail import send_mail
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse
from assessments.models import Assessment
import json
from django.utils import timezone
from django.db.utils import OperationalError
from utils.authenticators import generate_jwt_token
# import pandas lib as pd
import pandas as pd
import numpy as np
import os
# from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


# Create your views here.
# @api_view(['GET'])
# def getDashboardData(request):
#     try:
#         dataFrame = pd.read_excel('cairouniversity_march_known_nooutliers.xlsx')
#         print("hENAAAAA\n")
#         print(dataFrame)
#         #TODO:Uncomment trial & try again
#         # # Convert DataFrame to JSON
#         # # Loop over each column in the DataFrame
#         # for col in dataFrame.columns:
#         #     # Check if the column contains numeric data
#         #     if np.issubdtype(dataFrame[col].dtype, np.number):
#         #         # Replace out-of-range float values with a placeholder (np.nan)
#         #         dataFrame[col] = dataFrame[col].replace([np.inf, -np.inf], np.nan)
#         #         dataFrame[col] = dataFrame[col].astype(str)
#         # json_data = dataFrame.to_json(orient='records')
#         # Iterate over each row in the DataFrame
#         json_objects = []
#         for index, row in dataFrame.iterrows():
#             # Convert the row to a dictionary
#             row_dict = row.to_dict()
#             # Append the dictionary to the list
#             json_objects.append(row_dict)
#         return Response({'success': True, 'data': json_objects})
#         # return Response({'success': True,'message':'I\'m happy!'})
#     except FileNotFoundError as e:
#         # Return an error response if the file is not found
#         return Response({'success': False, 'message': f'File not found: {e}'}, status=404)
#     except OperationalError as e:
#         # Return an error response for database errors
#         return Response({'success': False, 'message': f'Database error: {e}'}, status=400)
#     except Exception as e:
#         # Return a generic error response for other exceptions
#         return Response({'success': False, 'message': f'An error occurred: {e}'}, status=500)

###Trial========================================
from django.http import JsonResponse
from rest_framework.decorators import api_view
import os

@api_view(['GET'])
def getDashboardData(request):

    # Read the Excel file
    try:
        data = pd.read_excel('cairouniversity_march_known_nooutliers.xlsx')
    except FileNotFoundError:
        return JsonResponse({"error": "File not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    # Replace NaN values with empty strings
    data = data.fillna('')
    # Convert the data to JSON format
    data_json = data.to_dict(orient='records')

    # Return the data as a JSON response
    return JsonResponse(data_json, safe=False)



#####
#======================================================================================
@api_view(['POST'])
def login(request):
    try:
        #TODO: return id , fisrat and lastname in the response -->DONE
        # Get request data
        email = request.data.get('email')
        password = request.data.get('password')

        # Check if both email and password are provided
        if email is None or password is None:
            return Response({'success': False, 'message': 'Please provide both email and password'}, status=400)

        # Check if email already exists in the database
        if DataScientist.objects.filter(email=email).exists():
            #compare password (methode 2)
            dataScientist=DataScientist.objects.get(email=email)
            print("My doctor object:\n",dataScientist)
            print("Database password:\n",dataScientist.password)
            print("User entered password:\n",password)

            print("Manual password checker:\n",type(password))
        
            # Check if user exists and password matches
            if dataScientist is None or not check_password(password, dataScientist.password):
                return Response({'success': False,'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
            else:
                #TODO:set an expiration date
                token=generate_jwt_token(dataScientist.id)
                print("My token: ",token)
                return Response({'success': True,"id":dataScientist.id,"firstName":dataScientist.firstName,"lastName":dataScientist.lastName,'token': token,'message': 'Login successful'})
        else:
            # Email does not exist in the database
            return Response({'success': False, 'error': 'Email does not exist'}, status=400)
    except OperationalError as e:
        # Return an error response for database errors
        return Response({'success': False, 'message': f'Database error: {e}'}, status=400)
    except Exception as e:
        # Return a generic error response for other exceptions
        return Response({'success': False, 'message': f'An error occurred: {e}'}, status=500)


#===========================END BRAND NEW LOGIN=====================

@api_view(['POST'])
def forgotPassword(request):
    try:    
        # Get request data
        email = request.data.get('email')
        password = request.data.get('password')

        # Check if both email and password are provided
        if email is None:
            return Response({'success': False, 'message': 'Please provide an email'}, status=400)
        if password is None:
            return Response({'success': False, 'message': 'Please provide a password'}, status=400)
        
        try:
            # Check if a Doctor object with the provided email exists
            dataScientist = DataScientist.objects.get(email=email)
            
            # Hash the new password
            hashed_password = make_password(password)
            
            # Update the password for the doctor object
            dataScientist.password = hashed_password
            
            # Save the doctor object to update the password in the database
            dataScientist.save()
            #TODO:Uncomment 
            # #send email to doctor
            # my_subject='Password notification for Baheya ALNM Website'
            # my_message='Your password is restored successfully. You can now login with your credentials.'
            # send_mail(
            #     subject=my_subject,
            #     message=my_message,
            #     recipient_list=[email],
            #     from_email=None,
            #     fail_silently=False)

            return Response({'success': True,'message': 'Password restored successfully.'})
        
        except ObjectDoesNotExist:
            # Handle the case where no Doctor object with the provided email exists
            return Response({'success': False, 'message': 'Email address not found.'}, status=404)
        
    except OperationalError as e:
        # Return an error response for database errors
        return Response({'success': False, 'message': f'Database error: {e}'}, status=400)
    except Exception as e:
        # Return a generic error response for other exceptions
        return Response({'success': False, 'message': f'An error occurred: {e}'}, status=500)


# #======================================================================================
# @api_view(['GET'])
# def export_assessments_to_excel(request):
#     try:
#         # Retrieve all Assessment records
#         assessments = Assessment.objects.all()

#         # Convert the QuerySet to a list of dictionaries
#         data_list = []
#         for assessment in assessments:
#             data_list.append({
#                 'id': assessment.id,
#                 'MRN': assessment.MRN,
#                 'status': assessment.status,
#                 'prediction': assessment.prediction,
#                 'ground_truth': assessment.ground_truth,
#                 'creation_date': assessment.creation_date,
#                 'medical_info': assessment.medical_info,
#             })

#         # Convert the list of dictionaries to a DataFrame
#         df = pd.DataFrame(data_list)

#         # Define the path to save the Excel file
#         file_path = 'assessments.xlsx'

#         # Use ExcelWriter to create the Excel file
#         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='Assessments')

#         return JsonResponse({'message': 'Assessments exported successfully', 'file_path': file_path}, status=200)
#     except Exception as e:
#         return JsonResponse({'error': str(e)}, status=500)
    

import pandas as pd
from django.http import JsonResponse
from rest_framework.decorators import api_view
from assessments.models import Assessment

# @api_view(['GET'])
# def export_assessments_to_excel(request):
#     try:
#         # Retrieve all Assessment records
#         assessments = Assessment.objects.all()

#         # Convert the QuerySet to a list of dictionaries
#         data_list = []
#         for assessment in assessments:
#             # Flatten the medical_info field into the main dictionary
#             assessment_dict = {
#                 'id': assessment.id,
#                 'MRN': assessment.MRN,
#                 'status': assessment.status,
#                 'prediction': assessment.prediction,
#                 'ground_truth': assessment.ground_truth,
               
#             }

#             # Include medical_info fields
#             medical_info = assessment.medical_info
#             if medical_info:
#                 for key, value in medical_info.items():
#                     assessment_dict[key] = value

#             data_list.append(assessment_dict)

#         # Convert the list of dictionaries to a DataFrame
#         df = pd.DataFrame(data_list)

#         # Define the path to save the Excel file
#         file_path = 'assessmentsMona.xlsx'

#         # Use ExcelWriter to create the Excel file
#         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='assessmentsMona')

#         return JsonResponse({'message': 'Assessments exported successfully', 'file_path': file_path}, status=200)
#     except Exception as e:
#         return JsonResponse({'error': str(e)}, status=500)


#TODO:Data Scientist will make a trial to see if these new rassessments addde to old excek are giving high accuracy and if so make status retrained(4)
#TODO:Noww(get assessments that have status (reviewed))


#TODO:Rga3y kol l json response ll assessments b tarteb l excel (Assessment dol)


@api_view(['GET'])
def export_assessments_to_excel(request):
    try:

        # Retrieve all Assessment records with status=1
        assessments = Assessment.objects.filter(status=1)


        # Convert the QuerySet to a list of dictionaries
        data_list = []
        for assessment in assessments:
            # Flatten the medical_info field into the main dictionary
            assessment_dict = {
                'ground_truth': assessment.ground_truth,
            }

            # Include medical_info fields
            medical_info = assessment.medical_info
            if medical_info:
                for key, value in medical_info.items():
                    assessment_dict[key] = value

            data_list.append(assessment_dict)

        
        # Convert the list of dictionaries to a DataFrame
        df = pd.DataFrame(data_list)

        # Specify the order of columns
        ordered_columns = [ 
                           'patient_first_bmi', 'patient_age', 'dm_result', 
                           'htn_result', 'vte_result', 'cvd_result', 
                           'Others', 'patient_family_history', 'patient_menopausal_state', 
                           'Hormonal_Contraception', 'patient_t', 
                           'patient_n', 'patient_size_cm', 'lymphovascular_invasion_result',
                           'patient_laterality','er_result','pr_result',
                           'her2_result','patient_ki67',
                           'patient_unilateral_bilateral','patient_site',
                           'patient_tumor_type','patient_grade','ground_truth']

        df = df[ordered_columns]

        # Define the path to save the Excel file
        file_path = 'cairouniversity_march_known_nooutliers copy.xlsx'
        sheet_name = 'Sheet1'

        # Check if the file exists
        if os.path.exists(file_path):
            # Load the workbook and the sheet
            book = openpyxl.load_workbook(file_path)

            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
            else:
                sheet = book.create_sheet(sheet_name)

            # Convert DataFrame to rows
            rows = dataframe_to_rows(df, index=False, header=False)

            # Append rows to the worksheet
            for row in rows:
                sheet.append(row)

        else:
            # If the file does not exist, create a new file
            df.to_excel(file_path, index=False, sheet_name=sheet_name)

        # Save the workbook
        book.save(file_path)

        return JsonResponse({'message': 'Assessments exported successfully', 'file_path': file_path}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
#================================================================================
#TODO: konty hatakhdelo dictionary{MRN:VALUR}
@api_view(['POST'])
def updateAssessmentToReviewed(request):
    try:
        data = request.data.get('MRNS', [])
        if not isinstance(data, list):
            return Response({
                'success': False,
                'message': 'Invalid input format. Expected a list of MRNs.'
            }, status=400)

        results = {}
        for MRN in data:
            try:
                assessment = Assessment.objects.get(MRN=MRN)
                assessment.status = 3
                assessment.save()
                results[MRN] = 'Assessment status updated successfully.'
            except Assessment.DoesNotExist:
                results[MRN] = 'This patient doesn\'t exist.'
            except Exception as e:
                results[MRN] = f'An error occurred: {e}'

        return Response({
            'success': True,
            'results': results
        })
    except OperationalError as e:
        return Response({'success': False, 'message': f'Database error: {e}'}, status=400)
    except Exception as e:
        return Response({'success': False, 'message': f'An error occurred: {e}'}, status=500)